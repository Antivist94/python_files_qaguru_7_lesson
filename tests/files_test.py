from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv
from script_os import ZIP_DIR


def test_check_pdf_file():
    with (ZipFile(ZIP_DIR) as zip_file):
        with zip_file.open('pdf_example.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            content = reader.pages[0]
            text = content.extract_text()
            assert 'Lorem ipsum \n' in text


def test_xlsx_file():
    with ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('xlsx_example.xlsx') as xlsx_file:
            wb = load_workbook(xlsx_file)
            sheet = wb.active

            age_values = []
            for row in sheet.iter_rows(min_row = 2, min_col = 6, max_col = 6, values_only = True):
                age_values.append(row[0])
            invalid_values = [str(value) for value in age_values if not str(value).isdigit()]

            assert not invalid_values, (f"Найдены недопустимые значения в столбце с возрастом: "
                                        f"{', '.join(invalid_values)}")


def test_csv_file():
    with ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('csv_example.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            head_row = csvreader[0]
            assert 'First Name' in head_row, 'В шапке таблицы должен быть столбец "First Name" '
            assert 'Last Name' in head_row, 'В шапке таблицы должен быть столбец "Last Name" '
            assert 'Gender' in head_row, 'В шапке таблицы должен быть столбец "Gender" '
            assert 'Country' in head_row, 'В шапке таблицы должен быть столбец "Country" '
            assert 'Age' in head_row, 'В шапке таблицы должен быть столбец "Age" '
            assert 'Date' in head_row, 'В шапке таблицы должен быть столбец "Date" '
            assert 'Id' in head_row, 'В шапке таблицы должен быть столбец "Id" '
